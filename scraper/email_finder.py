"""
email_finder.py — Phase 3: Find email addresses for exhibitor companies
------------------------------------------------------------------------
Strategy (in order of reliability):
  1. Scrape the company's own website — /contact, /about, homepage
  2. Scrape Google search results for  "company name" email contact
  3. Try common email pattern guessing (info@, contact@, sales@)
  4. Verify guessed addresses via MX record lookup (no SMTP ping needed)

Usage (standalone):
    python email_finder.py --input exhibitors.xlsx --out exhibitors_with_emails.xlsx

Usage (imported):
    from email_finder import EmailFinder
    finder = EmailFinder()
    results = await finder.enrich(rows)   # rows = list of exhibitor dicts

Requirements:
    pip install crawl4ai openpyxl playwright dnspython
"""

import asyncio
import json
import re
import socket
import argparse
from datetime import datetime
from urllib.parse import urljoin, urlparse

# Optional: dnspython for MX verification
try:
    import dns.resolver
    HAS_DNS = True
except ImportError:
    HAS_DNS = False

from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig, CacheMode
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, fills
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

# Pages to check on every company website (tried in order)
CONTACT_PATHS = [
    "/contact",
    "/contact-us",
    "/contact.html",
    "/about",
    "/about-us",
    "/impressum",        # German legal page — almost always has email
    "/legal",
    "/team",
    "/our-team",
    "/management",
    "/leadership",
    "/executive-team",
    "/management-team",
    "/board",
    "/board-of-directors",
    "/staff",
    "/key-personnel",
    "/company",
    "/info",
    "",                  # homepage last
]

# Roles to search for in web fallback
CONTACT_ROLES = [
    "CEO", "Founder", "Owner", "Director", "VP",
    "Managing Director", "President", "Chairman",
    "Head of Sales", "Sales Director", "Business Development",
    "CTO", "CFO", "COO", "CMO",
]

# Common generic email prefixes to guess
GUESS_PREFIXES = [
    "info",
    "contact",
    "sales",
    "hello",
    "mail",
    "office",
    "enquiries",
    "enquiry",
    "general",
    "admin",
]

# Regex that matches valid-looking emails (avoids image filenames etc.)
EMAIL_RE = re.compile(
    r"(?<![/\w])([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})(?!\.(?:png|jpg|gif|svg|webp|pdf|css|js))",
    re.VERBOSE,
)

# Domains to skip even if they appear in email addresses
JUNK_DOMAINS = {
    "example.com", "sentry.io", "wixpress.com", "squarespace.com",
    "amazonaws.com", "cloudfront.net", "googletagmanager.com",
    "facebook.com", "twitter.com", "linkedin.com", "instagram.com",
    "youtube.com", "tiktok.com", "pinterest.com",
    "w3.org", "schema.org", "ogp.me",
}


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def extract_emails_from_text(text: str) -> list[str]:
    """Pull all email-like strings from raw text, deduped and filtered."""
    found = EMAIL_RE.findall(text)
    cleaned = []
    seen = set()
    for e in found:
        e = e.lower().strip(".,;:'\"")
        domain = e.split("@")[-1]
        if domain in JUNK_DOMAINS:
            continue
        if e not in seen:
            seen.add(e)
            cleaned.append(e)
    return cleaned


def score_email(email: str, company_name: str = "", company_domain: str = "") -> int:
    """
    Score an email 0–100 for relevance.
    Higher = more likely to be the right contact.
    """
    score = 0
    local, domain = email.split("@")

    # Domain match with company website
    if company_domain and company_domain in domain:
        score += 50

    # Generic business prefixes are good
    if local in GUESS_PREFIXES:
        score += 20

    # Company name fragments in local part
    if company_name:
        name_parts = re.split(r"\W+", company_name.lower())
        for part in name_parts:
            if len(part) > 3 and part in local:
                score += 10

    # Penalise very long locals (likely automated/noreply)
    if len(local) > 20:
        score -= 10
    if "noreply" in local or "no-reply" in local or "donotreply" in local:
        score -= 30
    if "test" in local or "demo" in local or "example" in local:
        score -= 40

    return max(0, score)


def domain_from_url(url: str) -> str:
    try:
        return urlparse(url).netloc.replace("www.", "")
    except Exception:
        return ""


def build_contact_urls(website: str) -> list[str]:
    """Return ordered list of URLs to check for contact info."""
    if not website or not website.startswith("http"):
        return []
    base = website.rstrip("/")
    urls = []
    for path in CONTACT_PATHS:
        urls.append(base + path if path else base)
    return urls


async def verify_domain_mx(domain: str) -> bool:
    if not HAS_DNS:
        return True
    try:
        loop = asyncio.get_event_loop()
        resolver = dns.resolver.Resolver()
        resolver.timeout = 5
        resolver.lifetime = 5
        records = await loop.run_in_executor(
            None, lambda: resolver.resolve(domain, "MX")
        )
        return len(records) > 0
    except Exception:
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL FINDER
# ══════════════════════════════════════════════════════════════════════════════

class EmailFinder:

    def __init__(self, concurrency: int = 3, max_pages_per_company: int = 3,
                 verify_mx: bool = True, use_web_search: bool = True):
        """
        concurrency          — parallel companies to process
        max_pages_per_company — how many contact-path URLs to try before giving up
        verify_mx            — validate that email domains accept mail
        use_web_search       — also try Google search as a fallback
        """
        self.concurrency = concurrency
        self.max_pages = max_pages_per_company
        self.verify_mx = verify_mx
        self.use_web_search = use_web_search
        self._crawler = None
        self._browser_cfg = BrowserConfig(
            headless=True,
            verbose=False,
            extra_args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"],
        )

    # ── fetch a single URL and return raw text ───────────────────────────────

    async def _ensure_crawler(self):
        if self._crawler is None:
            self._crawler = await AsyncWebCrawler(config=self._browser_cfg).__aenter__()
        return self._crawler

    async def _fetch_text(self, url: str) -> str:
        run_cfg = CrawlerRunConfig(
            cache_mode=CacheMode.BYPASS,
            wait_for="body",
            page_timeout=20000,
            remove_overlay_elements=True,
            excluded_tags=["script", "style", "nav", "footer"],
            wait_for_images=False,
        )
        try:
            crawler = await self._ensure_crawler()
            result = await crawler.arun(url=url, config=run_cfg)
            if result.success:
                return (result.html or "") + "\n" + (result.markdown or "")
        except Exception:
            pass
        return ""

    async def close(self):
        if self._crawler is not None:
            await self._crawler.__aexit__(None, None, None)
            self._crawler = None

    # ── Strategy 1: scrape company website ───────────────────────────────────

    async def _emails_from_website(self, website: str, company_name: str) -> list[str]:
        """Try /contact, /about, /impressum, homepage in order."""
        urls_to_try = build_contact_urls(website)[:self.max_pages]
        company_domain = domain_from_url(website)
        all_emails = []

        for url in urls_to_try:
            text = await self._fetch_text(url)
            emails = extract_emails_from_text(text)
            if emails:
                # Score and prioritise domain-matching emails
                scored = sorted(
                    emails,
                    key=lambda e: score_email(e, company_name, company_domain),
                    reverse=True,
                )
                all_emails.extend(scored)
                # If we found a high-confidence email, stop early
                if score_email(scored[0], company_name, company_domain) >= 50:
                    break

        # Deduplicate preserving order
        seen, out = set(), []
        for e in all_emails:
            if e not in seen:
                seen.add(e)
                out.append(e)
        return out

    # ── Strategy 2: web search fallback ──────────────────────────────────────

    async def _emails_from_web_search(self, company_name: str, country: str = "") -> list[str]:
        """
        Build a Google search URL for the company + 'email contact' and scrape results.
        NOTE: Google blocks automated requests quickly; use DuckDuckGo HTML instead.
        """
        query = f"{company_name} email contact"
        if country:
            query += f" {country}"
        # DuckDuckGo lite (text-only, no JS required, not blocked easily)
        search_url = f"https://html.duckduckgo.com/html/?q={re.sub(r' ', '+', query)}"

        text = await self._fetch_text(search_url)
        return extract_emails_from_text(text)

    # ── Strategy 3: pattern guessing + MX verification ───────────────────────

    async def _guess_emails(self, website: str) -> list[str]:
        domain = domain_from_url(website)
        if not domain:
            return []
        guesses = [f"{prefix}@{domain}" for prefix in GUESS_PREFIXES]

        if self.verify_mx:
            mx_ok = await verify_domain_mx(domain)
            if not mx_ok:
                return []  # domain can't receive email — skip all guesses

        return guesses  # Return all guesses; caller can try SMTP or just keep top ones

    # ── Contact discovery ─────────────────────────────────────────────────────

    # Job title keywords to look for on team pages
    TITLE_KEYWORDS = [
        'ceo', 'cto', 'cfo', 'coo', 'cmo', 'cio', 'chief',
        'founder', 'co-founder', 'cofounder',
        'president', 'director', 'managing director', 'executive director',
        'vice president', 'vp', 'svp', 'avp', 'evp',
        'manager', 'senior manager', 'product manager', 'project manager',
        'head', 'head of', 'lead', 'team lead',
        'partner', 'managing partner', 'owner', 'principal',
        'sales', 'marketing', 'business development',
        'account manager', 'account executive',
        'chairman', 'chairperson', 'board member',
        'engineer', 'software engineer',
        'consultant', 'senior consultant', 'advisor',
        'operations', 'procurement',
    ]

    NAME_RE = re.compile(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\b')

    def _strip_html(self, text: str) -> str:
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'\s+', ' ', text)
        return text

    async def _search_contacts_by_role(self, company_name: str) -> list[dict]:
        """Search the web for people at a company by role (CEO, Founder, etc.)."""
        found = []
        seen = set()

        for role in CONTACT_ROLES:
            if len(found) >= 3:
                break
            query = f'"{company_name}" {role}'
            search_url = f"https://html.duckduckgo.com/html/?q={re.sub(r' ', '+', query)}"
            text = await self._fetch_text(search_url)
            if not text:
                continue

            # Look for names near role mentions
            clean = self._strip_html(text)
            lines = clean.split('\n')
            for i, line in enumerate(lines):
                line = line.strip()
                if not line or len(line) < 8:
                    continue
                if role.lower() not in line.lower():
                    continue

                names = self.NAME_RE.findall(line)
                for name in names:
                    if len(name.split()) < 2:
                        continue
                    key = f"{name}|{role}"
                    if key not in seen:
                        seen.add(key)
                        found.append({'name': name, 'title': role, 'email': ''})
                        if len(found) >= 3:
                            break

        return found

    async def _find_contacts_on_website(self, website: str, company_name: str) -> dict:
        """
        Visit company website pages and find people (name, title, email).
        Returns best contact dict: {contact_person, contact_title, email} or empty dict.
        """
        urls_to_try = build_contact_urls(website)[:self.max_pages]
        company_domain = domain_from_url(website)
        all_contacts = []
        seen_names = set()

        for url in urls_to_try:
            text = await self._fetch_text(url)
            if not text:
                continue
            text = self._strip_html(text)

            people = self._extract_people(text, url)
            for p in people:
                key = p['name'].lower()
                if key not in seen_names:
                    seen_names.add(key)
                    all_contacts.append(p)

        if not all_contacts and company_name:
            # Fallback: search web by role
            role_contacts = await self._search_contacts_by_role(company_name)
            for c in role_contacts:
                key = c['name'].lower()
                if key not in seen_names:
                    seen_names.add(key)
                    all_contacts.append(c)

        if not all_contacts:
            return {}

        # Score: prefer contacts with company-domain email, then senior titles
        for c in all_contacts:
            score = 0
            if c.get('email'):
                score += 40
                _, ed = c['email'].split('@')
                if ed == company_domain:
                    score += 60
            tl = c.get('title', '').lower()
            if any(kw in tl for kw in ['ceo', 'founder', 'president', 'director',
                                        'vp', 'chief', 'head', 'manager']):
                score += 30
            # Penalize if title is just the keyword (too vague)
            if len(tl) < 6:
                score -= 20
            c['_score'] = score

        all_contacts.sort(key=lambda c: c['_score'], reverse=True)
        best = all_contacts[0]

        result = {
            'contact_person': best['name'],
            'contact_title': best['title'],
        }

        # If best contact has a company-domain email, use it as primary
        if best.get('email'):
            _, ed = best['email'].split('@')
            if ed == company_domain:
                result['email'] = best['email']
                result['email_source'] = 'company_website'
                result['email_confidence'] = 'high'

        return result

    def _extract_people(self, text: str, page_url: str) -> list[dict]:
        """
        Extract people (name, title, email) from cleaned page text.
        Returns list of {name, title, email}.
        """
        people = []
        emails_with_pos = [(m.group(), m.start())
                           for m in EMAIL_RE.finditer(text.lower())]

        lines = text.split('\n')
        for i, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 5:
                continue

            title_lower = line.lower()
            matched_kw = next((kw for kw in self.TITLE_KEYWORDS
                               if kw in title_lower), None)
            if not matched_kw:
                continue

            name = None
            title = line.strip()

            # Pattern: Name — Title  or  Name | Title  or  Name, Title
            sep = re.search(r'\s*[–—\-|,•·:]\s*', line)
            if sep:
                before = line[:sep.start()].strip()
                after = line[sep.end():].strip()
                names_before = self.NAME_RE.findall(before)
                names_after = self.NAME_RE.findall(after)

                if names_before and len(before.split()) >= 2:
                    name = names_before[0]
                    title = after
                elif names_after and len(after.split()) >= 2:
                    title = before
                    name = names_after[0]

            if not name:
                names_in_line = self.NAME_RE.findall(line)
                if names_in_line:
                    name = names_in_line[0]

            if not name:
                for offset in [-1, 1]:
                    ni = i + offset
                    if 0 <= ni < len(lines):
                        nl = lines[ni].strip()
                        names_near = self.NAME_RE.findall(nl)
                        if names_near:
                            name = names_near[0]
                            break

            if not name:
                continue

            nearest_email = ''
            nearest_dist = 999999
            name_pos = text.find(name)
            if name_pos >= 0:
                for em, pos in emails_with_pos:
                    dist = abs(pos - name_pos)
                    if dist < nearest_dist:
                        nearest_dist = dist
                        nearest_email = em

            if nearest_email and nearest_dist > 500:
                nearest_email = ''

            people.append({
                'name': name,
                'title': title,
                'email': nearest_email,
            })

        return people

    # ── Main enrichment pipeline ──────────────────────────────────────────────

    async def _find_for_one(self, row: dict) -> dict:
        """Run all strategies for one exhibitor and annotate the row."""
        company   = row.get("company_name", "")
        website   = row.get("website", "")
        country   = row.get("country", "")
        known_email = row.get("email", "")

        # Already has a real email — skip
        if known_email and "@" in known_email:
            row["email_source"] = "scraped_from_show_site"
            row["email_confidence"] = "high"
            return row

        all_emails: list[str] = []
        source = ""

        # ── Strategy 1: company website ──────────────────────────────────────
        if website:
            print(f"    🌐  {company[:40]:<40}  website search…")
            found = await self._emails_from_website(website, company)
            if found:
                all_emails = found
                source = "company_website"

        # ── Strategy 2: web search (if website found nothing) ─────────────────
        if not all_emails and self.use_web_search and company:
            print(f"    🔍  {company[:40]:<40}  web search…")
            found = await self._emails_from_web_search(company, country)
            if found:
                all_emails = found
                source = "web_search"

        # ── Strategy 3: pattern guessing (last resort) ────────────────────────
        if not all_emails and website:
            print(f"    💡  {company[:40]:<40}  pattern guess…")
            found = await self._guess_emails(website)
            if found:
                all_emails = found
                source = "guessed_pattern"

        # ── Contact discovery: find people on company website ────────────────
        if website:
            contacts = await self._find_contacts_on_website(website, company)
            if contacts.get('contact_person'):
                print(f"    👤  {company[:40]:<40}  found contact: {contacts['contact_person']} ({contacts.get('contact_title','?')})")
                row['contact_person'] = contacts['contact_person']
                row['contact_title']  = contacts['contact_title']
                # If discovery found a company-domain email, prefer it
                if contacts.get('email') and contacts.get('email_source') == 'company_website':
                    row['email']          = contacts['email']
                    row['email_source']   = 'company_website'
                    row['email_confidence'] = 'high'

        # ── Pick best email ───────────────────────────────────────────────────
        if all_emails:
            company_domain = domain_from_url(website) if website else ""
            scored = sorted(
                all_emails,
                key=lambda e: score_email(e, company, company_domain),
                reverse=True,
            )
            row["email"]          = scored[0]
            row["email_alts"]     = "; ".join(scored[1:4])  # up to 3 alternatives
            row["email_source"]   = source
            row["email_confidence"] = (
                "high"   if source == "company_website" else
                "medium" if source == "web_search"      else
                "low"
            )
        else:
            row["email"]          = ""
            row["email_alts"]     = ""
            row["email_source"]   = "not_found"
            row["email_confidence"] = ""

        return row

    async def enrich(self, rows: list[dict]) -> list[dict]:
        """Enrich a list of exhibitor dicts with email data. Returns enriched list."""
        semaphore = asyncio.Semaphore(self.concurrency)
        total = len(rows)
        done = 0

        async def bounded(row):
            nonlocal done
            async with semaphore:
                try:
                    result = await asyncio.wait_for(self._find_for_one(row), timeout=60)
                except asyncio.TimeoutError:
                    result = row
                done += 1
                found = "✓" if result.get("email") else "–"
                contact = "👤" if result.get("contact_person") else "  "
                print(f"  [{done:>3}/{total}] {found}{contact}  {result.get('company_name','')[:50]}")
                return result

        try:
            tasks = [bounded(r) for r in rows]
            enriched = await asyncio.gather(*tasks, return_exceptions=True)
            return [r for r in enriched if isinstance(r, dict)]
        finally:
            await self.close()


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL UPDATE — adds email columns to existing exhibitors.xlsx
# ══════════════════════════════════════════════════════════════════════════════

EMAIL_COLS = [
    ("Email",              28),
    ("Alt Emails",         38),
    ("Email Source",       20),
    ("Confidence",         14),
]

CONF_COLORS = {
    "high":   "C6EFCE",  # green
    "medium": "FFEB9C",  # yellow
    "low":    "FFCCCC",  # red/pink
    "":       "F2F2F2",
}

def update_excel_with_emails(rows: list[dict], input_path: str, output_path: str):
    """Load existing xlsx, add/update email columns, save to new path."""
    try:
        wb = load_workbook(input_path)
        ws = wb["Exhibitors"]
    except Exception:
        # Build fresh workbook if input doesn't exist
        from openpyxl import Workbook as WB
        wb = WB()
        ws = wb.active
        ws.title = "Exhibitors"

    # Find last used column
    last_col = ws.max_column

    # Check if email columns already exist
    headers = [ws.cell(1, c).value for c in range(1, last_col + 1)]
    if "Email" not in headers:
        # Append email column headers
        for i, (header, width) in enumerate(EMAIL_COLS):
            col = last_col + 1 + i
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            cell.fill      = PatternFill("solid", start_color="1A3C5E")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        email_col_start = last_col + 1
    else:
        email_col_start = headers.index("Email") + 1

    # Build lookup: company_name → row number
    name_col = next(
        (i + 1 for i, h in enumerate(headers) if h and "company" in str(h).lower()),
        1
    )
    name_to_row = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, name_col).value
        if val:
            name_to_row[str(val).strip().lower()] = r

    # Write email data
    thin = Side(style="thin", color="C0D9E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for item in rows:
        name = str(item.get("company_name", "")).strip().lower()
        row_num = name_to_row.get(name)
        if not row_num:
            continue

        email      = item.get("email", "")
        alts       = item.get("email_alts", "")
        source     = item.get("email_source", "")
        confidence = item.get("email_confidence", "")
        conf_color = CONF_COLORS.get(confidence, "F2F2F2")

        values = [email, alts, source, confidence]
        for i, val in enumerate(values):
            col = email_col_start + i
            c = ws.cell(row=row_num, column=col, value=val)
            c.font      = Font(name="Calibri", size=10,
                               color="0563C1" if i == 0 and "@" in str(val) else "000000",
                               underline="single" if i == 0 and "@" in str(val) else None)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = border
            if i == 3:  # confidence column — colour coded
                c.fill = PatternFill("solid", start_color=conf_color)
            else:
                alt_row_fill = "E8F1F8" if row_num % 2 == 0 else "FFFFFF"
                c.fill = PatternFill("solid", start_color=alt_row_fill)

    # Stats sheet update
    if "Summary" in wb.sheetnames:
        ws2 = wb["Summary"]
        n = ws.max_row
        next_row = ws2.max_row + 2
        ws2.cell(next_row,     1, "Emails Found").font    = Font(bold=True, name="Calibri", size=10)
        ws2.cell(next_row,     2, f"=COUNTA(Exhibitors!{get_column_letter(email_col_start)}2:{get_column_letter(email_col_start)}{n})").font = Font(name="Calibri", size=10)
        ws2.cell(next_row + 1, 1, "Email Coverage %").font = Font(bold=True, name="Calibri", size=10)
        ws2.cell(next_row + 1, 2, f"=IFERROR({get_column_letter(email_col_start + 0)}{next_row}/B3,0)").number_format = "0%"
        ws2.cell(next_row + 1, 2).font = Font(name="Calibri", size=10)

    wb.save(output_path)
    found = sum(1 for r in rows if r.get("email"))
    print(f"\n✅  Excel saved → {output_path}")
    print(f"   Emails found: {found}/{len(rows)}  ({found*100//max(len(rows),1)}%)")


# ══════════════════════════════════════════════════════════════════════════════
#  STANDALONE CLI
# ══════════════════════════════════════════════════════════════════════════════

def load_rows_from_excel(path: str) -> list[dict]:
    """Read exhibitor rows from an existing xlsx produced by exhibitor_scraper.py"""
    wb = load_workbook(path, data_only=True)
    ws = wb["Exhibitors"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h:
                row[h.lower().replace(" ", "_").replace("/", "_")] = ws.cell(r, c).value
        if row.get("company_name"):
            rows.append(row)
    return rows


async def main():
    p = argparse.ArgumentParser(description="Add email addresses to exhibitor Excel")
    p.add_argument("--input",       required=True,  help="Input xlsx from exhibitor_scraper.py")
    p.add_argument("--out",         default="",     help="Output xlsx (default: overwrites input with _emails suffix)")
    p.add_argument("--no-search",   action="store_true", help="Disable web search fallback (website scraping only)")
    p.add_argument("--no-mx",       action="store_true", help="Skip MX record verification")
    p.add_argument("--concurrency", type=int, default=3, help="Parallel requests (default: 3)")
    p.add_argument("--limit",       type=int, default=0, help="Only process first N companies (0 = all)")
    args = p.parse_args()

    output = args.out or args.input.replace(".xlsx", "_emails.xlsx")

    print(f"\n📥  Loading exhibitors from {args.input}…")
    rows = load_rows_from_excel(args.input)
    if args.limit:
        rows = rows[:args.limit]
    print(f"   {len(rows)} companies loaded\n")

    finder = EmailFinder(
        concurrency=args.concurrency,
        verify_mx=not args.no_mx,
        use_web_search=not args.no_search,
    )

    print("🔍  Finding emails…\n")
    enriched = await finder.enrich(rows)

    print(f"\n💾  Writing to Excel…")
    update_excel_with_emails(enriched, args.input, output)


if __name__ == "__main__":
    asyncio.run(main())
