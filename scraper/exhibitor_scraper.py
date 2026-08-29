"""
exhibitor_scraper.py — Scrape EXHIBITOR lists from tradeshow websites
-----------------------------------------------------------------------
Phase 1 — collect the exhibitor LIST, following pagination to the end
Phase 2 — follow each exhibitor's detail link for deeper info (optional)

The work is split across three modules:
  list_crawler.py  — walks every page of a listing (pagination, load-more)
  extractors.py    — picks the best container selector and pulls fields
  llm_extract.py   — model-based extraction for pages selectors can't read

Usage:
    python exhibitor_scraper.py --url "https://example-show.com/exhibitors"
    python exhibitor_scraper.py --url "https://example-show.com/exhibitors" --deep
    python exhibitor_scraper.py --urls urls.txt --out my_exhibitors.xlsx

Requirements:
    pip install playwright beautifulsoup4 lxml openpyxl aiohttp
    playwright install chromium
"""

import asyncio
import json
import os
import re
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from urllib.parse import urljoin

from site_configs import get_config_for_domain
from list_crawler import ListCrawler
from extractors import extract_from_pages, extract_detail_page, dedupe
from llm_extract import extract_exhibitors


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

class ExhibitorScraper:

    def __init__(self, use_llm: bool = False, deep: bool = False,
                 llm_provider: str = "", max_detail_pages: int = 0,
                 max_list_pages: int = 100):
        self.use_llm = use_llm
        self.deep = deep
        self.llm_provider = llm_provider or os.environ.get("LLM_PROVIDER", self._default_provider())
        # 0 means "no cap". The old default of 50 silently deep-scraped only the
        # first 50 exhibitors of a show — 4% of a 1,200-exhibitor directory —
        # with nothing in the output to say the rest had been skipped.
        self.max_detail_pages = max_detail_pages or int(os.environ.get("MAX_DETAIL_PAGES", "0"))
        self.max_list_pages = max_list_pages
        self.results: list[dict] = []

    @staticmethod
    def _default_provider() -> str:
        if os.environ.get("OPENAI_API_KEY"):
            return "openai/gpt-4o-mini"
        return "ollama/llama3"


    # ── Phase 1: scrape exhibitor list ───────────────────────────────────────

    async def scrape_list_page(self, url: str) -> list[dict]:
        """
        Scrape an exhibitor directory, following pagination to the end.

        Order of preference:
          1. A site-specific config (an internal API is always the best source)
          2. Generic multi-page collection + CSS extraction
          3. LLM extraction over the collected text, when --llm is on
        """
        domain_config = get_config_for_domain(url)
        if domain_config:
            if domain_config.get("type") == "api":
                rows = await self._scrape_via_api(url, domain_config["api_config"])
                if rows:
                    return self._finalise(rows, url)
                print("  ! Site API returned nothing — falling back to the generic crawler")
            elif domain_config.get("type") == "playwright":
                from site_configs import get_playwright_scraper
                scraper_fn = get_playwright_scraper(domain_config.get("playwright_scraper"))
                if scraper_fn:
                    rows = await scraper_fn(url)
                    if rows:
                        return self._finalise(rows, url)
                    print("  ! Site scraper returned nothing — falling back to the generic crawler")

        # ── Collect every page of the listing ────────────────────────────
        crawler = ListCrawler(max_pages=self.max_list_pages, verbose=True)
        pages_html = await crawler.collect(url)
        if not pages_html:
            print(f"  ✗ Could not load {url}")
            return []

        # ── CSS extraction across all collected pages ────────────────────
        rows, selector = extract_from_pages(pages_html, url)
        if rows:
            print(f"  ✓ Extracted {len(rows)} exhibitors using selector: {selector}")

        # ── LLM fallback, when the selectors found little or nothing ─────
        # A directory page that yields under 5 rows is almost always a miss,
        # not a genuinely tiny show.
        if len(rows) < 5 and self.use_llm:
            print(f"  ℹ CSS extraction found {len(rows)} rows — trying the LLM…")
            llm_rows = await self._llm_rows_from_pages(pages_html)
            if len(llm_rows) > len(rows):
                print(f"  ✓ LLM extraction found {len(llm_rows)} exhibitors")
                rows = llm_rows

        if not rows:
            print(f"  ✗ No exhibitors extracted from {url}")
            if not self.use_llm:
                print("    Try re-running with LLM Extraction enabled, or add a")
                print("    site config in scraper/site_configs.py for this domain.")
            return []

        return self._finalise(rows, url)

    def _finalise(self, rows: list[dict], url: str) -> list[dict]:
        """Normalise URLs and stamp provenance onto every row."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for r in rows:
            r.setdefault("source_url", url)
            r.setdefault("scraped_at", now)
            # urljoin handles root-relative paths and query strings correctly;
            # the old string concatenation mangled both.
            for field in ("detail_url", "website"):
                value = (r.get(field) or "").strip()
                if value and not value.startswith(("http://", "https://")):
                    r[field] = urljoin(url, value)
        rows = dedupe(rows)
        print(f"  ✓ List page total: {len(rows)} exhibitors  ←  {url}")
        return rows

    async def _llm_rows_from_pages(self, pages_html: list[str]) -> list[dict]:
        """Strip each collected page to text and run the LLM over it."""
        from bs4 import BeautifulSoup

        texts = []
        for html in pages_html:
            soup = BeautifulSoup(html, "lxml")
            for tag in soup(["script", "style", "nav", "footer", "header", "svg"]):
                tag.decompose()
            texts.append(soup.get_text("\n"))

        rows: list[dict] = []
        for text in texts:
            rows.extend(await extract_exhibitors(text))
        return dedupe(rows)

    async def _scrape_via_api(self, url: str, cfg: dict) -> list[dict]:
        """
        Scrape exhibitors via a backend JSON/XML API (site_configs.py).
        Uses Playwright to load the init page for session cookies, then calls
        the API endpoint with pagination.
        """
        import xml.etree.ElementTree as ET
        from playwright.async_api import async_playwright

        rows = []
        page_size = cfg.get("page_size", 200)
        endpoint = cfg["endpoint"]
        base_params = dict(cfg.get("base_params", {}))

        try:
            async with async_playwright() as pw:
                browser = await pw.chromium.launch(
                    headless=True,
                    args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"],
                )
                page = await browser.new_page()

                init_url = cfg.get("init_page", url)
                await page.goto(init_url, wait_until="networkidle", timeout=30000)

                start = 0
                total_expected = None
                while True:
                    params = dict(base_params)
                    params["numresultrows"] = str(page_size)
                    params["startresultrow"] = str(start)

                    js = "const fd = new URLSearchParams();\n"
                    for k, v in params.items():
                        js += f'fd.append("{k}", "{v}");\n'
                    js += f"""
                        const r = await fetch("{endpoint}", {{
                            method: "POST",
                            headers: {{"Content-Type": "application/x-www-form-urlencoded"}},
                            body: fd.toString()
                        }});
                        return await r.text();
                    """

                    result_xml = await page.evaluate(f"async () => {{ {js} }}")

                    root = ET.fromstring(result_xml)
                    entities = root.find(".//entities")
                    if entities is None:
                        break

                    count = int(entities.get("count", 0))
                    if total_expected is None:
                        total_expected = count
                        print(f"  ℹ API reports {total_expected} total exhibitors")

                    for org in entities.findall("organization"):
                        lead = {"source_url": url, "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M")}
                        field_map = cfg.get("field_map", {})
                        for field_name, mapping in field_map.items():
                            if "path" in mapping:
                                elem = org.find(mapping["path"])
                                if elem is not None:
                                    if mapping.get("text"):
                                        val = (elem.text or "").strip()
                                    else:
                                        val = elem.attrib.get(mapping.get("attr", ""), "")
                                else:
                                    val = ""
                            elif "attr" in mapping:
                                val = org.attrib.get(mapping["attr"], "")
                            else:
                                val = ""
                            lead[field_name] = val
                        rows.append(lead)

                    has_more = entities.get("hasMore")
                    if has_more == "false":
                        break
                    if not has_more and len(entities.findall("organization")) < page_size:
                        break
                    start += page_size

                await browser.close()

        except Exception as e:
            print(f"  ✗ API scraping failed: {e}")
            import traceback
            traceback.print_exc()

        return rows

    # ── Phase 2 (optional): deep-scrape individual company profiles ──────────

    async def scrape_detail_pages(self, rows: list[dict]) -> list[dict]:
        """
        STEP 4 (optional, activated by --deep flag).
        For each exhibitor that has a detail_url, fetches the profile page
        and merges richer data back into the row.
        """
        with_urls = [r for r in rows if r.get("detail_url")]
        to_fetch = with_urls[:self.max_detail_pages] if self.max_detail_pages else with_urls
        if not to_fetch:
            print("  ℹ No detail URLs found — skipping deep scrape.")
            return rows

        print(f"\n  📄  Deep-scraping {len(to_fetch)} of {len(with_urls)} exhibitor profile pages…")
        if len(to_fetch) < len(with_urls):
            print(f"  ⚠ Capped at {self.max_detail_pages} by MAX_DETAIL_PAGES — "
                  f"{len(with_urls) - len(to_fetch)} profiles will be skipped.")
        semaphore = asyncio.Semaphore(3)  # polite concurrency

        # Detail pages are read with Playwright + the same field heuristics as
        # the list page. The old version always used an LLM strategy regardless
        # of --llm, so with no API key every profile silently failed.
        from playwright.async_api import async_playwright
        from list_crawler import BROWSER_ARGS, USER_AGENT

        done = 0
        enriched_count = 0

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=BROWSER_ARGS)
            context = await browser.new_context(user_agent=USER_AGENT)

            async def fetch_one(row: dict) -> dict:
                nonlocal done, enriched_count
                async with semaphore:
                    page = await context.new_page()
                    try:
                        await page.goto(row["detail_url"], wait_until="domcontentloaded", timeout=30000)
                        await page.wait_for_timeout(600)
                        html = await page.content()
                        detail = extract_detail_page(html, row["detail_url"])
                        if detail:
                            enriched_count += 1
                        for k, v in detail.items():
                            if v and not row.get(k):
                                row[k] = v
                    except Exception as e:
                        print(f"    ! detail page failed for {row.get('company_name','?')[:40]}: {type(e).__name__}")
                    finally:
                        await page.close()
                        done += 1
                        if done % 25 == 0 or done == len(to_fetch):
                            print(f"    [{done}/{len(to_fetch)}] profiles fetched")
                    return row

            enriched = await asyncio.gather(*[fetch_one(r) for r in to_fetch],
                                            return_exceptions=True)
            await browser.close()

        print(f"  ✓ Deep scrape enriched {enriched_count}/{len(to_fetch)} profiles")

        enriched_map = {r["detail_url"]: r for r in enriched
                        if isinstance(r, dict) and r.get("detail_url")}
        return [enriched_map.get(r.get("detail_url"), r) for r in rows]

    # ── Main pipeline ────────────────────────────────────────────────────────

    async def run(self, urls: list[str]) -> list[dict]:
        all_rows = []
        for n, url in enumerate(urls, 1):
            print(f"\n[{n}/{len(urls)}] {url}")
            rows = await self.scrape_list_page(url)
            if self.deep and rows:
                rows = await self.scrape_detail_pages(rows)
            all_rows.extend(rows)

        # Shows often list the same exhibitor across several URLs (per-hall
        # pages, A-Z splits); collapse them rather than writing duplicates.
        before = len(all_rows)
        all_rows = dedupe(all_rows)
        if before != len(all_rows):
            print(f"\n  ℹ Merged {before - len(all_rows)} duplicate exhibitors across URLs")

        self.results = all_rows
        return all_rows



# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("Company Name",      28),
    ("Booth / Stand",     14),
    ("Hall / Pavilion",   16),
    ("Country",           16),
    ("City",              16),
    ("Category / Sector", 24),
    ("Products / Services",28),
    ("Description",       45),
    ("Website",           32),
    ("Email",             26),
    ("Phone",             18),
    ("Contact Person",    22),
    ("LinkedIn",          30),
    ("Detail Profile URL",36),
    ("Source URL",        36),
    ("Scraped At",        18),
]

FIELD_MAP = [
    "company_name", "booth_number", "hall", "country", "city",
    "category", "products", "description",
    "website", "email", "phone", "contact_person",
    "social_linkedin", "detail_url", "source_url", "scraped_at",
]

HDR_FILL   = PatternFill("solid", start_color="1A3C5E")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ALT_FILL   = PatternFill("solid", start_color="E8F1F8")
NORM_FILL  = PatternFill("solid", start_color="FFFFFF")
LINK_FONT  = Font(color="0563C1", underline="single", name="Calibri", size=10)
BODY_FONT  = Font(name="Calibri", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="C0D9E8")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
URL_FIELDS = {"website", "social_linkedin", "detail_url", "source_url"}


def export_to_excel(rows: list[dict], path: str = "exhibitors.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Exhibitors"
    ws.freeze_panes = "A2"

    # Header
    for ci, (header, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, CENTER, BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    # Data
    for ri, item in enumerate(rows, 2):
        fill = ALT_FILL if ri % 2 == 0 else NORM_FILL
        for ci, field in enumerate(FIELD_MAP, 1):
            val = item.get(field) or ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.border, c.fill = BORDER, fill
            if field in URL_FIELDS and val and str(val).startswith("http"):
                c.hyperlink = val
                c.font, c.alignment = LINK_FONT, LEFT
            else:
                c.font, c.alignment = BODY_FONT, LEFT
        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    n = len(rows) + 1
    summary_data = [
        ("Total Exhibitors",    f"=COUNTA(Exhibitors!A2:A{n})"),
        ("Unique Countries",    f"=IFERROR(SUMPRODUCT(1/COUNTIF(Exhibitors!D2:D{n},Exhibitors!D2:D{n})),0)"),
        ("Unique Categories",   f"=IFERROR(SUMPRODUCT(1/COUNTIF(Exhibitors!F2:F{n},Exhibitors!F2:F{n})),0)"),
        ("With Website",        f"=COUNTA(Exhibitors!I2:I{n})"),
        ("With Email",          f"=COUNTA(Exhibitors!J2:J{n})"),
        ("Generated",           datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    ws2["A1"] = "Exhibitor Data Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri", color="1A3C5E")
    for i, (label, val) in enumerate(summary_data, 3):
        ws2.cell(i, 1, label).font = Font(bold=True, name="Calibri", size=10)
        ws2.cell(i, 2, val).font   = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    wb.save(path)
    print(f"\n✅  Saved → {path}  ({len(rows)} exhibitors)")


def export_to_csv(rows: list[dict], path: str):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_MAP, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅  CSV  → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    p = argparse.ArgumentParser(description="Scrape exhibitor lists → Excel")
    p.add_argument("--url",    help="Single exhibitor-list URL")
    p.add_argument("--urls",   help="Text file with one URL per line")
    p.add_argument("--llm",    action="store_true", help="Use LLM extraction (needs OPENAI_API_KEY)")
    p.add_argument("--deep",   action="store_true", help="Also scrape individual exhibitor profile pages")
    p.add_argument("--emails", action="store_true", help="Phase 3: hunt for email addresses on the web")
    p.add_argument("--max-list-pages", dest="max_list_pages", type=int, default=100)
    p.add_argument("--max-detail-pages", dest="max_detail_pages", type=int, default=0)
    p.add_argument("--out",    default="exhibitors.xlsx")
    p.add_argument("--csv",    action="store_true")
    args = p.parse_args()

    if args.url:
        urls = [args.url]
    elif args.urls:
        urls = [l.strip() for l in Path(args.urls).read_text().splitlines()
                if l.strip() and not l.startswith("#")]
    else:
        print("Provide --url or --urls. Example:\n  python exhibitor_scraper.py --url https://myshow.com/exhibitors")
        return

    print(f"\n🔍  Scraping {len(urls)} URL(s) | LLM={'on' if args.llm else 'off'} | Deep={'on' if args.deep else 'off'} | Emails={'on' if args.emails else 'off'}\n")
    scraper = ExhibitorScraper(use_llm=args.llm, deep=args.deep,
                               max_detail_pages=args.max_detail_pages,
                               max_list_pages=args.max_list_pages)
    rows = await scraper.run(urls)

    if not rows:
        print("⚠️  No exhibitors found. See STEP 2 in the guide to inspect selectors.")
        return

    # ── Phase 3: email enrichment ─────────────────────────────────────────────
    if args.emails:
        try:
            from email_finder import EmailFinder, update_excel_with_emails
            print(f"\n📧  Phase 3 — hunting emails for {len(rows)} companies…\n")
            finder = EmailFinder(concurrency=3, verify_mx=True, use_web_search=True)
            rows = await finder.enrich(rows)
            # Save Excel with email columns added
            export_to_excel(rows, args.out)
            # Re-open and inject colour-coded confidence columns
            out_emails = args.out.replace(".xlsx", "_emails.xlsx")
            update_excel_with_emails(rows, args.out, out_emails)
            print(f"\n✅  Final file with emails → {out_emails}")
        except ImportError:
            print("⚠️  email_finder.py not found next to this script. Skipping email phase.")
            export_to_excel(rows, args.out)
    else:
        export_to_excel(rows, args.out)

    if args.csv:
        export_to_csv(rows, args.out.replace(".xlsx", ".csv"))


if __name__ == "__main__":
    asyncio.run(main())
